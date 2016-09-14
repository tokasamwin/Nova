from openpyxl import load_workbook
import numpy as np
import pylab as pl
from amigo import geom
import scipy as sp
from scipy.interpolate import interp1d
from collections import OrderedDict
from itertools import cycle,count
from scipy.linalg import norm 
from matplotlib.collections import PolyCollection
import pandas as pd
import seaborn as sns
rc = {'figure.figsize':[7,7*12/9],'savefig.dpi':100, # 
      'savefig.jpeg_quality':100,'savefig.pad_inches':0.1,
      'lines.linewidth':1.5}
sns.set(context='talk',style='white',font='sans-serif',palette='Set2',
        font_scale=7/8,rc=rc)
        
def get_label(label,label_array,force=False,part=None):
    if label != None:
        label_array.append(label.replace(' ','_'))
        flag = True
    else:
        if force:
            label_array.append(part)
            if len(label_array) == 1:
                flag = True
            else:
                if label_array[-2] != label_array[-1]:
                    flag = True
                else:
                    flag = False
        else:
            flag = False
    return flag
            
def add_segment(lines,k):
    segment = '{:1.0f}'.format(next(k))
    lines[segment] = {'r':np.array([]),'z':np.array([])}
    return segment
    
def append_value(lines,segment,r,z):
    lines[segment]['r'] = np.append(lines[segment]['r'],r)
    lines[segment]['z'] = np.append(lines[segment]['z'],z)
    
def cutcorners(r,z):
    r,z = geom.pointloop(r,z,ref='min')  # form loop from closest neighbour
    n = len(r)-1
    dR = np.array([r[1:]-r[:-1],z[1:]-z[:-1]])
    dR = np.array([np.gradient(r),np.gradient(z)])
    dL = norm(dR,axis=0)
    k,kflag = count(0),False
    lines = OrderedDict()
    segment = add_segment(lines,k)
    for i in range(n):
        kink = np.arccos(np.dot(dR[:,i]/dL[i],
                                dR[:,i+1]/dL[i+1]))*180/np.pi
        append_value(lines,segment,r[i+1],z[i+1])
        if abs(kink) > 40 and kflag == False:  # angle, deg
            segment = add_segment(lines,k)
            append_value(lines,segment,r[i+1],z[i+1])
            kflag = True
        else:
            kflag = False
    segments = list(lines.keys())
    l = np.zeros(len(segments))
    for i,seg in enumerate(segments):
        l[i] = len(lines[seg]['r'])
    seg = np.argsort(l)[-2:]  # select loops (in/out)
    rmax = np.zeros(len(seg))
    for i,s in enumerate(seg):
        rmax[i] = np.max(lines[segments[s]]['r'])
    seg = seg[np.argsort(rmax)]
    lines_sort = OrderedDict()
    for s in seg[:2]:
        lines_sort[segments[s]] = lines[segments[s]]
    return lines_sort
    
def cluster_points(r,z):
    R,Z = geom.pointloop(r,z)
    dX = norm([r[1:]-r[:-1],z[1:]-z[:-1]],axis=0)
    dx_median = sp.median(dX)
    cluster,i = OrderedDict(),count(0)
    for r,z in zip(R,Z):
        dx = []
        for cl in cluster:
            rc,zc = cluster[cl]['r'],cluster[cl]['z']
            dx.append(np.min(norm([r-rc,z-zc],
                                            axis=0)))
        if len(dx) == 0 or np.min(dx) > 2*dx_median:  # new
            cl = 'group{:1.0f}'.format(next(i))
            cluster[cl] = {}
            cluster[cl] = {'r':[r],'z':[z]}
        else:
            icl = np.argmin(dx)
            cl = list(cluster.keys())[icl]
            cluster[cl]['r'] = np.append(cluster[cl]['r'],r)
            cluster[cl]['z'] = np.append(cluster[cl]['z'],z)
    for cl in cluster:    
        r,z = cluster[cl]['r'],cluster[cl]['z']
        dx = norm([r[:1]-r[:-1],z[:1]-z[:-1]],axis=0)
        imax = np.argmax(dx)+1
        r = np.append(r[imax:],r[:imax])
        z = np.append(z[imax:],z[:imax])
        cluster[cl]['r'],cluster[cl]['z'] = r,z
    return cluster
    
def set_figure():
    pl.axis('off')
    pl.axis('equal')
    
class DEMO(object):
    
    def __init__(self):
        self.filename = 'DEMO1_Reference_Design_-_2015_April_(_EU_2MG46D_v1_0'
        self.read(self.filename)
        self.process()
        
    def read(self,filename):
        wb = load_workbook(filename='./referance/'+filename+
                           '.xlsx',read_only=True,
                   data_only=True)
        ws = wb[wb.get_sheet_names()[0]]
        self.parts = OrderedDict()  # component parts    
        part,loop = [],[]
        for row in ws.columns:
            new_part = get_label(row[0].value,part)
            if new_part:
                self.parts[part[-1]] = OrderedDict()
            if len(part) == 0:
                continue
            new_loop = get_label(row[1].value,loop,force=new_part,part=part[-1])
            p,l = part[-1],loop[-1]
            if new_loop:
                self.parts[p][l] = OrderedDict()
            
            comp,start = row[2].value,2
            while comp == None:
                start += 1
                comp = row[start].value
            if comp == 'Rad':  # replace confusing 'Rad' label
                comp = 'r'
                
            self.parts[p][l][comp] = np.zeros(len(row)-1)
            for i,r in enumerate(row[start+1:]):
                try:
                    self.parts[p][l][comp][i] = 1e-3*float(r.value)  # m
                except:
                    break
            self.parts[p][l][comp] = self.parts[p][l][comp][:i]
        
    def process(self):
        for part in self.parts:
            x = {'in':{'r':np.array([]),'z':np.array([])},
                 'out':{'r':np.array([]),'z':np.array([])},
                 'ports':{'r':np.array([]),'z':np.array([])},
                 'r':np.array([]),'z':np.array([])}
                 
            for loop,side in zip(self.parts[part],['out','in','ports']):
                r,z = geom.read_loop(self.parts[part],loop)
                x[side]['r'],x[side]['z'] = r,z 
                
            if part in ['TF_Coil','Vessel','Blanket']:
                if side != 'out':
                    x = geom.polyloop(x)
                else:
                    x['r'],x['z'] = geom.pointloop(x['out']['r'],x['out']['z'])
                    lines = cutcorners(x['r'],x['z'])  # select halfs
                    for seg,side in zip(lines,['in','out']):
                        x[side] = {'r':lines[seg]['r'],'z':lines[seg]['z']}
            for key in x:
                self.parts[part][key] = x[key]
                                  
    def get_ports(self,plot=True):
        x = self.parts['Vessel']['ports']
        clusters = cluster_points(x['r'],x['z'])
        self.port = OrderedDict()
        for i,cl in enumerate(clusters): 
            r,z = clusters[cl]['r'],clusters[cl]['z']
            switch = r.max()-r.min() < 0.5*(z.max()-z.min())  
            if switch:  # rotate coordinates
                x,y = np.copy(z),np.copy(r)
            else:
                x,y = np.copy(r),np.copy(z)
            index = np.argsort(x)
            x,y,r,z = x[index],y[index],r[index],z[index]
            M = np.array([np.ones(len(x)),x]).T  # linear least-squares fit
            a = sp.linalg.lstsq(M,y)[0]
            fit = a[0]+a[1]*x
            if switch:
                r_fit,z_fit = fit,z
            else:
                r_fit,z_fit = r,fit
            n_hat = np.array([-(z_fit[-1]-z_fit[0]),
                              r_fit[-1]-r_fit[0]])
            n_hat /= norm(n_hat)
            n = len(r)
            count = {'left':0,'right':0}
            p = 'P{:1.0f}'.format(i)
            self.port[p] = {'left':{'r':np.zeros(n),'z':np.zeros(n)},
                            'right':{'r':np.zeros(n),'z':np.zeros(n)}}
            for r_,z_ in zip(r,z):   
                for dot,side in zip([1,-1],['left','right']):
                    if dot*np.dot([r_-r_fit[0],z_-z_fit[0]],n_hat) > 0:
                        self.port[p][side]['r'][count[side]] = r_
                        self.port[p][side]['z'][count[side]] = z_
                        count[side] += 1
            for side in ['left','right']:
                for var in ['r','z']:  # trim
                    n = count[side]
                    self.port[p][side][var] = self.port[p][side][var][:n]
                if plot:
                    pl.plot(self.port[p][side]['r'],self.port[p][side]['z'],
                            color=0.65*np.ones(3))
                        
    def get_limiters(self,plot=True):
        x = self.parts['Plasma']['out']
        self.limiter = OrderedDict()
        clusters = cluster_points(x['r'],x['z'])
        for i,cl in enumerate(clusters): 
            r,z = clusters[cl]['r'],clusters[cl]['z']
            self.limiter['L{:1.0f}'.format(i)] = {'r':r,'z':z}
            if plot:
                pl.plot(r,z,color=0.5*np.ones(3))
                
    def get_fw(self,plot=True):
        rbl = self.parts['Blanket']['in']['r']  # read blanket
        zbl = self.parts['Blanket']['in']['z']
        zmin = np.zeros(len(self.limiter))  # select divertor limiter
        for i,limiter in enumerate(self.limiter):
            zmin[i] = np.min(self.limiter[limiter]['z'])
        imin = np.argmin(zmin)
        div = list(self.limiter.keys())[imin]
        rdiv = self.limiter[div]['r']  # set divertor profile
        zdiv = self.limiter[div]['z']
        cut = np.zeros(2,dtype=int)  # cut and join divertor
        for i,j in enumerate([0,-1]):
            cut[i] = np.argmin(norm([rbl[j]-rdiv,zbl[j]-zdiv],axis=0))
        cut = np.sort(cut)
        r = np.append(rbl,rdiv[cut[0]:cut[1]])
        z = np.append(zbl,zdiv[cut[0]:cut[1]])
        r,z = np.append(r,r[0]),np.append(z,z[0])
        r,z = geom.rzSLine(r,z)
        r,z = r[::-1],z[::-1]  # flip
        self.fw = {'r':r,'z':z}
        if plot:
            pl.plot(r,z,color=0.5*np.ones(3))
          
    def fill_loops(self):
        color = cycle(sns.color_palette('Set2',5))
        for part in self.parts:
            try:
                geom.polyfill(self.parts[part]['r'],
                              self.parts[part]['z'],color=next(color))       
            except:
                pass

    def plot(self):
        for part in self.parts:
            for loop in self.parts[part]:
                pl.plot(self.parts[part][loop]['r'],
                        self.parts[part][loop]['z'],'.',markersize=5.0)
                        
    def write(self):
        filename = 'DEMO1_sorted'
        wb = load_workbook(filename='./referance/'+self.filename+'.xlsx')
                           
        ws = wb[wb.get_sheet_names()[0]]
        
        part = iter(self.parts)
        component = cycle(['r','z'])
        for row in ws.columns:
            if row[0].value is not None:
                pt = next(part)
                row[0].value = pt
                loop = iter(['out','in','ports'])
            if row[1].value is not None:
                lp = next(loop)
                row[1].value = lp
            if row[2].value is not None:
                cp = next(component)
                row[2].value = cp
            '''    
            for i,r in enumerate(row[3:]):
                if r.value is not None:
                    print(pt,lp,cp,i)
                    r.value = self.parts[pt][lp][cp][i]
            '''    
                
            '''
            new_loop = get_label(row[1].value,loop,force=new_part,part=part[-1])
            p,l = part[-1],loop[-1]
            if new_loop:
                self.parts[p][l] = OrderedDict()
            
            comp,start = row[2].value,2
            while comp == None:
                start += 1
                comp = row[start].value
            if comp == 'Rad':  # replace confusing 'Rad' label
                comp = 'r'
                
            self.parts[p][l][comp] = np.zeros(len(row)-1)
            for i,r in enumerate(row[start+1:]):
                try:
                    self.parts[p][l][comp][i] = 1e-3*float(r.value)  # m
                except:
                    break
            self.parts[p][l][comp] = self.parts[p][l][comp][:i]
            '''
        wb.save('./referance/'+filename+'.xlsx')
        
if __name__ is '__main__':          
        demo = DEMO()
        
        demo.fill_loops()
        demo.get_limiters()
        demo.get_ports()
        demo.get_fw()
        
        #demo.write()

        #demo.plot()
        set_figure()
        
        
        print(pd.DataFrame(demo.parts))

        from nova.config import Setup
        from nova.streamfunction import SF
        from nova.radial_build import RB
        from nova.elliptic import EQ
        from nova.coils import PF,TF
        from nova.inverse import INV
        
        import seaborn as sns
        rc = {'figure.figsize':[8*12/16,8],'savefig.dpi':120, # 
              'savefig.jpeg_quality':100,'savefig.pad_inches':0.1,
              'lines.linewidth':2}
        sns.set(context='talk',style='white',font='sans-serif',palette='Set2',
                font_scale=7/8,rc=rc)
        
        config = 'DEMO_SN'
        setup = Setup(config)
        sf = SF(setup.filename)
        
        pf = PF(sf.eqdsk)
        eq = EQ(sf,pf,limit=[4.5,14.5,-8,8],n=5e3) 
        rb = RB(setup,sf)
        
        r = demo.parts['Vessel']['out']['r']  # set vessel boundary
        z = demo.parts['Vessel']['out']['z']

        r,z = geom.offset(r,z,0.2)  # 200mm offset from vessel 
        r,z = geom.rzSLine(r,z,npoints=20)
        rb.loop = geom.Loop(r,z)
        
        tf = TF(shape={'vessel':rb.loop,'pf':pf,'sf':sf,
                       'fit':False,'setup':setup,
                       'coil_type':'S',
                       'config':config})  # ,'config':config
        
        #tf.coil.plot()
        tf.load_coil()
        x = tf.coil.draw()
        tf.get_coil_loops(x['r'],x['z'],profile='in')
        tf.fill()
        
        xnorm,bnorm = tf.set_oppvar()

        for var in tf.coil.xo:
            tf.coil.xo[var]['xnorm'] = (tf.coil.xo[var]['value']-
                                        tf.coil.xo[var]['lb'])/\
                                        (tf.coil.xo[var]['ub']-
                                        tf.coil.xo[var]['lb'])
        data = pd.DataFrame(tf.coil.xo).T
        data.reset_index(level=0,inplace=True)
        print(data)
        

        pl.figure(figsize=(8,4))
        sns.set_color_codes("muted")
        sns.barplot(x='xnorm',y='index',data=data,color="b")
        sns.despine(bottom=True)
        pl.ylabel('')
        ax = pl.gca()
        ax.get_xaxis().set_visible(False)
        patch = ax.patches

        eps = 1e-2
        values = [tf.coil.xo[var]['value'] for var in tf.coil.xo]
        xnorms = [tf.coil.xo[var]['xnorm'] for var in tf.coil.xo]
        for p,value,xnorm in zip(patch,values,xnorms):
            
            x = p.get_width()
            print(xnorm,x)
            y = p.get_y() 
            if xnorm < eps or xnorm > 1-eps:
                color = 'r'
            else:
                color = 'k'
            ax.text(x,y,' {:1.3f}'.format(value),ha='left',va='top',
                    size='small',color=color)
        pl.plot(0.5*np.ones(2),np.sort(ax.get_ylim()),'--',color=0.5*np.ones(3),
                zorder=0,lw=1)
        pl.plot(np.ones(2),np.sort(ax.get_ylim()),'-',color=0.5*np.ones(3),
                zorder=0,lw=2)
        pl.xlim([0,1])



        '''
        referance = Dcoil()
        x = referance.draw()
        referance.plot()
        self.get_coil_loops(x['r'],x['z'],profile='in')
        coil = {'Rcl':self.Rcl,'Zcl':self.Zcl,
                'nTF':self.nTF,'Iturn':self.Iturn}
        rp = ripple(plasma={'config':'SN'},coil=coil)
        print('ref ripple',rp.get_ripple())
        '''  
        
        '''
        rb.Rb,rb.Zb = demo.fw['r'],demo.fw['z']  # set radial build first wall
        rb.trim_sol(Nsol=3,update=False,plot=False)
        inv = INV(sf,pf,eq)
        
        inv.fix_boundary_psi(N=23,alpha=1-1e-4,factor=1)  # add boundary points
        inv.fix_boundary_feild(N=23,alpha=1-1e-4,factor=1)  # add boundary points
        inv.add_null(factor=1,point=sf.Xpoint)
        for leg in ['inner','outer']:
            point = [sf.legs[leg]['R'][0][-1],sf.legs[leg]['Z'][0][-1]]
            #if leg == 'inner':
            #    point[1] += 0.4
            inv.add_alpha(1,factor=1,point=point)
        #inv.fix['z'] += 0.3
        inv.plot_fix()
        '''
        
        '''
        Lpf = inv.grid_PF(nPF=5)
        Lcs = inv.grid_CS(nCS=5)
        Lo = np.append(Lpf,Lcs)
        inv.eq.coils()  # re-grid
        inv.update_coils()
        '''

        '''
        inv.initalise_plasma = True
        inv.set_plasma()
        inv.swing_fix(0)
        inv.solve_slsqp()
        '''
        
        '''
        pf.plot(coils=pf.coil,label=False,plasma=False,current=False) 
        
        #eq.gen_opp()
        #eq.run()
        rb.trim_sol(Nsol=3,update=True,plot=True)
        sf.contour(Nlevel=31,plot_vac=False)
        '''
