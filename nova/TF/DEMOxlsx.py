from openpyxl import load_workbook
import numpy as np
import pylab as pl
from amigo import geom
from nova.config import trim_dir
import scipy as sp
from scipy.interpolate import interp1d
from collections import OrderedDict
from itertools import cycle,count
from scipy.linalg import norm 
from matplotlib.collections import PolyCollection
import seaborn as sns
rc = {'figure.figsize':[7,7*12/9],'savefig.dpi':100, # 
      'savefig.jpeg_quality':100,'savefig.pad_inches':0.1,
      'lines.linewidth':1.5}
sns.set(context='talk',style='white',font='sans-serif',palette='Set2',
        font_scale=7/8,rc=rc)
color = sns.color_palette('Paired',12)
ic = count(0)
        
def get_label(label,label_array,force=False,part=None):
    if label != None:
        label_array.append(label.replace(' ','_'))
        flag = True
    else:
        if force:
            label_array.append(part)
            if len(label_array) == 1:
                flag = True
            else:
                if label_array[-2] != label_array[-1]:
                    flag = True
                else:
                    flag = False
        else:
            flag = False
    return flag
            
def add_segment(lines,k):
    segment = '{:1.0f}'.format(next(k))
    lines[segment] = {'r':np.array([]),'z':np.array([])}
    return segment
    
def append_value(lines,segment,r,z):
    lines[segment]['r'] = np.append(lines[segment]['r'],r)
    lines[segment]['z'] = np.append(lines[segment]['z'],z)
    
def cutcorners(r,z):
    r,z = geom.pointloop(r,z,ref='min')  # form loop from closest neighbour
    n = len(r)-1
    dR = np.array([r[1:]-r[:-1],z[1:]-z[:-1]])
    dR = np.array([np.gradient(r),np.gradient(z)])
    dL = norm(dR,axis=0)
    k,kflag = count(0),False
    lines = OrderedDict()
    segment = add_segment(lines,k)
    for i in range(n):
        kink = np.arccos(np.dot(dR[:,i]/dL[i],
                                dR[:,i+1]/dL[i+1]))*180/np.pi
        append_value(lines,segment,r[i+1],z[i+1])
        if abs(kink) > 40 and kflag == False:  # angle, deg
            segment = add_segment(lines,k)
            append_value(lines,segment,r[i+1],z[i+1])
            kflag = True
        else:
            kflag = False
    segments = list(lines.keys())
    l = np.zeros(len(segments))
    for i,seg in enumerate(segments):
        l[i] = len(lines[seg]['r'])
    seg = np.argsort(l)[-2:]  # select loops (in/out)
    rmax = np.zeros(len(seg))
    for i,s in enumerate(seg):
        rmax[i] = np.max(lines[segments[s]]['r'])
    seg = seg[np.argsort(rmax)]
    lines_sort = OrderedDict()
    for s in seg[:2]:
        lines_sort[segments[s]] = lines[segments[s]]
    return lines_sort
    
def cluster_points(r,z):
    R,Z = geom.pointloop(r,z)
    dX = norm([r[1:]-r[:-1],z[1:]-z[:-1]],axis=0)
    dx_median = sp.median(dX)
    cluster,i = OrderedDict(),count(0)
    for r,z in zip(R,Z):
        dx = []
        for cl in cluster:
            rc,zc = cluster[cl]['r'],cluster[cl]['z']
            dx.append(np.min(norm([r-rc,z-zc],axis=0)))
        if len(dx) == 0 or np.min(dx) > 2*dx_median:  # new
            cl = 'group{:1.0f}'.format(next(i))
            cluster[cl] = {}
            cluster[cl] = {'r':[r],'z':[z]}
        else:
            icl = np.argmin(dx)
            cl = list(cluster.keys())[icl]
            cluster[cl]['r'] = np.append(cluster[cl]['r'],r)
            cluster[cl]['z'] = np.append(cluster[cl]['z'],z)
    for cl in cluster:    
        r,z = cluster[cl]['r'],cluster[cl]['z']
        dx = norm([r[:1]-r[:-1],z[:1]-z[:-1]],axis=0)
        imax = np.argmax(dx)+1
        r = np.append(r[imax:],r[:imax])
        z = np.append(z[imax:],z[:imax])
        cluster[cl]['r'],cluster[cl]['z'] = r,z
    return cluster
    
def set_figure():
    pl.axis('off')
    pl.axis('equal')
    
class DEMO(object):
    
    def __init__(self):
        self.filename = 'DEMO1_Reference_Design_-_2015_April_(_EU_2MG46D_v1_0'
        self.read(self.filename)
        self.process()
        self.get_limiters()
        self.get_ports()
        self.get_fw()
        
    def read(self,filename):
        ref = trim_dir('../../TF/referance/')
        wb = load_workbook(filename=ref+filename+
                           '.xlsx',read_only=True,
                   data_only=True)
        ws = wb[wb.get_sheet_names()[0]]
        self.parts = OrderedDict()  # component parts    
        part,loop = [],[]
        for row in ws.columns:
            new_part = get_label(row[0].value,part)
            if new_part:
                self.parts[part[-1]] = OrderedDict()
            if len(part) == 0:
                continue
            new_loop = get_label(row[1].value,loop,force=new_part,part=part[-1])
            p,l = part[-1],loop[-1]
            if new_loop:
                self.parts[p][l] = OrderedDict()
            
            comp,start = row[2].value,2
            while comp == None:
                start += 1
                comp = row[start].value
            if comp == 'Rad':  # replace confusing 'Rad' label
                comp = 'r'
                
            self.parts[p][l][comp] = np.zeros(len(row)-1)
            for i,r in enumerate(row[start+1:]):
                try:
                    self.parts[p][l][comp][i] = 1e-3*float(r.value)  # m
                except:
                    break
            self.parts[p][l][comp] = self.parts[p][l][comp][:i]
        
    def process(self):
        for part in self.parts:
            x = {'in':{'r':np.array([]),'z':np.array([])},
                 'out':{'r':np.array([]),'z':np.array([])},
                 'ports':{'r':np.array([]),'z':np.array([])},
                 'r':np.array([]),'z':np.array([])}
                 
            for loop,side in zip(self.parts[part],['out','in','ports']):
                r,z = geom.read_loop(self.parts[part],loop)
                x[side]['r'],x[side]['z'] = r,z 
                
            if part in ['TF_Coil','Vessel','Blanket']:
                if side != 'out':
                    x['r'],x['z'] = geom.polyloop(x['in'],x['out'])
                else:
                    x['r'],x['z'] = geom.pointloop(x['out']['r'],x['out']['z'])
                    lines = cutcorners(x['r'],x['z'])  # select halfs
                    for seg,side in zip(lines,['in','out']):
                        x[side] = {'r':lines[seg]['r'],'z':lines[seg]['z']}
            for key in x:
                self.parts[part][key] = x[key]
                                  
    def get_ports(self,plot=False):
        x = self.parts['Vessel']['ports']
        clusters = cluster_points(x['r'],x['z'])
        port = OrderedDict()
        for i,cl in enumerate(clusters): 
            r,z = clusters[cl]['r'],clusters[cl]['z']
            switch = r.max()-r.min() < 0.5*(z.max()-z.min())  
            if switch:  # rotate coordinates
                x,y = np.copy(z),np.copy(r)
            else:
                x,y = np.copy(r),np.copy(z)
            index = np.argsort(x)
            x,y,r,z = x[index],y[index],r[index],z[index]
            M = np.array([np.ones(len(x)),x]).T  # linear least-squares fit
            a = sp.linalg.lstsq(M,y)[0]
            fit = a[0]+a[1]*x
            if switch:
                r_fit,z_fit = fit,z
            else:
                r_fit,z_fit = r,fit
            n_hat = np.array([-(z_fit[-1]-z_fit[0]),
                              r_fit[-1]-r_fit[0]])
            n_hat /= norm(n_hat)
            n = len(r)
            count = {'left':0,'right':0}
            p = 'P{:1.0f}'.format(i)
            port[p] = {'left':{'r':np.zeros(n),'z':np.zeros(n)},
                            'right':{'r':np.zeros(n),'z':np.zeros(n)}}
            for r_,z_ in zip(r,z):   
                for dot,side in zip([1,-1],['left','right']):
                    if dot*np.dot([r_-r_fit[0],z_-z_fit[0]],n_hat) > 0:
                        port[p][side]['r'][count[side]] = r_
                        port[p][side]['z'][count[side]] = z_
                        count[side] += 1
            for side in ['left','right']:
                for var in ['r','z']:  # trim
                    n = count[side]
                    port[p][side][var] = port[p][side][var][:n]
        ro = np.mean(self.parts['Blanket']['r'])
        zo = np.mean(self.parts['Blanket']['z'])
        theta = np.zeros(len(port))
        for i,p in enumerate(port):
            theta[i] = np.arctan2(port[p]['left']['z'][0]-zo,
                                  port[p]['left']['r'][0]-ro)
        index = list(np.argsort(theta))
        pkey = list(port.keys())
        self.port = OrderedDict()
        for i,indx in enumerate(index):
            psort = 'P{:1.0f}'.format(i)
            self.port[psort] = port[pkey[indx]] 
        
        if plot:
            self.plot_ports()
 
    def plot_ports(self):
        for p in self.port:
            for s in self.port[p]:
                if s == 'left':
                    c=color[8]
                else:
                    c=color[8]
                pl.plot(self.port[p][s]['r'],self.port[p][s]['z'],
                        zorder=3,color=c,lw=1)
                    
    def get_limiters(self,plot=False):
        x = self.parts['Plasma']['out']
        self.limiter = OrderedDict()
        clusters = cluster_points(x['r'],x['z'])
        for i,cl in enumerate(clusters): 
            r,z = clusters[cl]['r'],clusters[cl]['z']
            self.limiter['L{:1.0f}'.format(i)] = {'r':r,'z':z}
            if plot:
                pl.plot(r,z,color=0.5*np.ones(3))
                
    def plot_limiter(self):
        pl.plot(self.limiter['L3']['r'],
                self.limiter['L3']['z'],color=0.6*np.ones(3)) 
                
    def get_fw(self,plot=False):
        rbl = self.parts['Blanket']['in']['r']  # read blanket
        zbl = self.parts['Blanket']['in']['z']
        zmin = np.zeros(len(self.limiter))  # select divertor limiter
        for i,limiter in enumerate(self.limiter):
            zmin[i] = np.min(self.limiter[limiter]['z'])
        imin = np.argmin(zmin)
        div = list(self.limiter.keys())[imin]
        rdiv = self.limiter[div]['r']  # set divertor profile
        zdiv = self.limiter[div]['z']
        cut = np.zeros(2,dtype=int)  # cut and join divertor
        for i,j in enumerate([0,-1]):
            cut[i] = np.argmin(norm([rbl[j]-rdiv,zbl[j]-zdiv],axis=0))
        cut = np.sort(cut)
        r = np.append(rbl,rdiv[cut[0]:cut[1]])
        z = np.append(zbl,zdiv[cut[0]:cut[1]])
        r,z = np.append(r,r[0]),np.append(z,z[0])
        r,z = geom.rzSLine(r,z)
        r,z = r[::-1],z[::-1]  # flip
        self.fw = {'r':r,'z':z}
        if plot:
            pl.plot(r,z,color=0.5*np.ones(3))
          
    def fill_loops(self):
        for part in self.parts:
            try:
                geom.polyfill(self.parts[part]['r'],
                              self.parts[part]['z'],color=color(next(ic)))       
            except:
                pass
        set_figure()
        
    def fill_part(self,part):
        if part == 'TF_Coil':
            cindex = len(self.parts.keys())
        else:
            cindex = list(self.parts.keys()).index(part)-1
        geom.polyfill(self.parts[part]['r'],self.parts[part]['z'],
                      color=color[cindex])   

    def plot(self):
        for part in self.parts:
            for loop in self.parts[part]:
                try:
                    pl.plot(self.parts[part][loop]['r'],
                            self.parts[part][loop]['z'],'.',markersize=5.0)
                except:
                    pass
                        
    def write(self):
        filename = 'DEMO1_sorted'
        wb = load_workbook(filename='./referance/'+self.filename+'.xlsx')
                           
        ws = wb[wb.get_sheet_names()[0]]
        
        part = iter(self.parts)
        component = cycle(['r','z'])
        for row in ws.columns:
            if row[0].value is not None:
                pt = next(part)
                row[0].value = pt
                loop = iter(['out','in','ports'])
            if row[1].value is not None:
                lp = next(loop)
                row[1].value = lp
            if row[2].value is not None:
                cp = next(component)
                row[2].value = cp
            '''    
            for i,r in enumerate(row[3:]):
                if r.value is not None:
                    print(pt,lp,cp,i)
                    r.value = self.parts[pt][lp][cp][i]
            '''    
                
            '''
            new_loop = get_label(row[1].value,loop,force=new_part,part=part[-1])
            p,l = part[-1],loop[-1]
            if new_loop:
                self.parts[p][l] = OrderedDict()
            
            comp,start = row[2].value,2
            while comp == None:
                start += 1
                comp = row[start].value
            if comp == 'Rad':  # replace confusing 'Rad' label
                comp = 'r'
                
            self.parts[p][l][comp] = np.zeros(len(row)-1)
            for i,r in enumerate(row[start+1:]):
                try:
                    self.parts[p][l][comp][i] = 1e-3*float(r.value)  # m
                except:
                    break
            self.parts[p][l][comp] = self.parts[p][l][comp][:i]
            '''
        wb.save('./referance/'+filename+'.xlsx')
        
if __name__ is '__main__':          
        demo = DEMO()
        
        demo.fill_loops()

        demo.plot()
        demo.plot_ports()
        #demo.write()
        
        
        